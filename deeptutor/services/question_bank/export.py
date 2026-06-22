"""Excel export for question-bank (notebook) entries."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook

from deeptutor.services.question_bank.filters import is_failed_generation_entry

EXPORT_HEADERS = ("题目", "选项A", "选项B", "选项C", "选项D", "答案", "解析")
_OPTION_KEYS = ("A", "B", "C", "D")


def _option_value(options: dict[str, Any] | None, key: str) -> str:
    if not isinstance(options, dict):
        return ""
    return str(options.get(key) or options.get(key.lower()) or "").strip()


def entries_to_xlsx_bytes(entries: list[dict[str, Any]]) -> bytes:
    """Build an ``.xlsx`` workbook from notebook entry dicts."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "题目"
    sheet.append(list(EXPORT_HEADERS))
    for entry in entries:
        if is_failed_generation_entry(entry):
            continue
        options = entry.get("options")
        sheet.append(
            [
                str(entry.get("question") or "").strip(),
                _option_value(options, "A"),
                _option_value(options, "B"),
                _option_value(options, "C"),
                _option_value(options, "D"),
                str(entry.get("correct_answer") or "").strip(),
                str(entry.get("explanation") or "").strip(),
            ]
        )
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def parse_xlsx_rows(data: bytes) -> list[tuple[str, ...]]:
    """Read exported rows back — used by tests."""
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.active
    return [tuple(str(cell or "") for cell in row) for row in sheet.iter_rows(values_only=True)]


__all__ = ["EXPORT_HEADERS", "entries_to_xlsx_bytes", "parse_xlsx_rows"]
